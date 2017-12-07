import os
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict
import configparser
import lucernex

# Create the full path to the Templates folder
current_working_directory = os.path.dirname(__file__)

# Read the reports.conf file and load into memory
reports = configparser.ConfigParser()
reports.read('reports.conf')


def build_field_dict_from_template(report):
    table = reports[report]["table"]
    file = get_template_file(report)

    wb = load_workbook(filename=file)
    ws = wb.get_sheet_by_name(table)
    columns = ws.max_column

    field_dict = {}

    # Fields start in column 5, start with 4 since each loop increments
    index_number = 4

    # iterate through each column and grab anything after the first period
    for column in ws.iter_cols(min_row=2, min_col=5, max_col=columns, max_row=2):
        index_number += 1
        for cell in column:
            field_name = cell.value.split(".")
            field_dict[index_number] = field_name[1]

    return field_dict


def get_template_file(report):
    templates_folder = os.path.join(current_working_directory, 'Templates')
    return os.path.join(templates_folder, reports[report]["template"])


def create_report_file_name(name):
    # Only creating Excel reports with this program
    name += ".xlsx"
    reports_folder = os.path.join(current_working_directory, 'Reports')
    return os.path.join(reports_folder, name)


def write_to_excel(data, report, name=None):
    table = reports[report]["table"]
    template_file = get_template_file(report)
    fields_dict = build_field_dict_from_template(report)

    wb = load_workbook(filename=template_file)
    ws = wb.get_sheet_by_name(table)

    row_number = 3

    for item in data:
        row_number += 1

        # Each row needs Payment in the first column
        ws.cell(column=2, row=row_number, value=table)
        ws.cell(column=4, row=row_number, value=item)

        for field in fields_dict:
            field_name = fields_dict[field]
            try:
                if isinstance(data[item][field_name], dict):
                    text_value = data[item][field_name]['Name']
                else:
                    text_value = data[item][field_name]
            except KeyError:
                # Do not populate anything for empty keys
                text_value = ""

            ws.cell(column=field, row=row_number, value=text_value)

    # Set the name of the report to the report if left blank
    name = report if name is None else name

    wb.save(create_report_file_name(name))


def run_report(report, firm_name="Default", environment="TRAIN", name=None):
    api = reports[report]["api"]
    table = reports[report]["table"]

    if api == "fiql":
        data = lucernex.fiql_get(table, firm_name, environment, reports[report]["fiql"])
    else:
        pass

    write_to_excel(data, report, name)
